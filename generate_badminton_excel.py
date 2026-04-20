#!/usr/bin/env python3
"""
Tạo file Excel quản lý chi phí & công nợ nhóm cầu lông (7 TV cố định + GL).
Công nợ dùng SUMPRODUCT+SUMIF+INDIRECT — tương thích Microsoft Excel và LibreOffice Calc
(không dùng REDUCE/LAMBDA/FILTER của Excel 365).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

MASTER = "Công nợ & Quy định"
GUIDE = "Hướng dẫn"
TEMPLATE = "Mẫu_phiên"
SAMPLE_DAY = "2026-04-20"

# Ô tham số trên sheet master (cột H) — cùng hàng với nhãn cột G
PARAM_EXTRA_COURT_ROW = 6
PARAM_GL_M_ROW = 7
PARAM_GL_F_ROW = 8
PARAM_THR_HIGH_ROW = 9
PARAM_THR_LOW_ROW = 10

# Danh sách TV cố định: cột B, bắt đầu hàng
FIXED_NAMES_START_ROW = 52

thin = Side(style="thin", color="444444")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(color="FFFFFF", bold=True)
title_font = Font(bold=True, size=14)
sub_font = Font(bold=True, size=11)

# Cột A16:A201 = vùng tên sheet phiên (khớp bảng tblPhien; sửa cả đây nếu đổi vùng trong file)
PHIEN_SHEET_NAMES_REF = "$A$16:$A$201"


def luy_ke_cell_formula(row: int) -> str:
    """Lũy kế: SUMPRODUCT+SUMIF+INDIRECT; TRIM tránh khoảng trắng ẩn làm INDIRECT lỗi."""
    r = PHIEN_SHEET_NAMES_REF
    return (
        f'=IF($A{row}="","",'
        f"SUMPRODUCT((LEN(TRIM({r}))>0)*"
        f'IFERROR(SUMIF(INDIRECT("\'"&TRIM({r})&"\'!$B$20:$B$35"),'
        f'$A{row},INDIRECT("\'"&TRIM({r})&"\'!$F$20:$F$35")),0)))'
    )


def so_du_cell_formula(row: int) -> str:
    """Số dư = cột C − cột D (tham chiếu ô, tránh lỗi [@Tên] trên một số bản Calc)."""
    return f'=IF($A{row}="","",IFERROR(C{row}-D{row},0))'


def add_master(wb: Workbook) -> None:
    ws = wb.create_sheet(MASTER, 0)

    ws["A1"] = "Quản lý công nợ & quy định — Nhóm cầu lông"
    ws["A1"].font = title_font
    ws.merge_cells("A1:J1")

    ws["A3"] = "Quy định tính tiền (tóm tắt)"
    ws["A3"].font = sub_font
    rules = (
        "1) Tiền sân thêm: nhập số sân thêm; hệ thống nhân với đơn giá ở bảng Tham số.\n"
        "2) Tổng chi buổi = Tiền sân cố định (buổi) + (Số sân thêm × đơn giá) + Tiền cầu.\n"
        "3) Chỉ đếm người có mặt: TV cố định cột «Có mặt» = 1; GL có tên ở bảng GL.\n"
        "4) Nếu tổng người có mặt > ngưỡng «Trên ngưỡng GL» (mặc định 12): "
        "mỗi GL Nam đóng «GL Nam», mỗi GL Nữ đóng «GL Nữ»; "
        "phần còn lại của tổng chi chia đều cho các TV cố định có mặt (GL không chia phần sân/cầu trong nhánh này).\n"
        "5) Nếu tổng người ≤ ngưỡng trên và ≥ ngưỡng dưới (mặc định 8): chia đều tổng chi cho mọi người có mặt.\n"
        "6) Nếu tổng người < ngưỡng dưới: vẫn chia đều tổng chi cho người có mặt; "
        "khuyến nghị không đặt sân thêm (kiểm tra tay số sân thêm).\n"
        "7) Cột «Số dư» = «Lũy kế phải đóng» − «Đã thanh toán». Dương = còn nợ, Âm = dư / được hoàn."
    )
    ws["A4"] = rules
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F12")
    ws.row_dimensions[4].height = 180

    # Tham số
    ws["G3"] = "Tham số (sửa tại đây)"
    ws["G3"].font = sub_font
    ws["G5"] = "Mục"
    ws["H5"] = "Giá trị (VNĐ / người)"
    for c in ("G5", "H5"):
        ws[c].fill = hdr_fill
        ws[c].font = hdr_font
        ws[c].border = border_all

    params = [
        ("Đơn giá 1 sân thêm", 300_000),
        ("Phí GL Nam (khi > ngưỡng)", 80_000),
        ("Phí GL Nữ (khi > ngưỡng)", 60_000),
        ("Ngưỡng trên (áp phí GL riêng)", 12),
        ("Ngưỡng dưới (cảnh báo ít người)", 8),
    ]
    for i, (label, val) in enumerate(params, start=PARAM_EXTRA_COURT_ROW):
        ws.cell(row=i, column=7, value=label).border = border_all
        cell = ws.cell(row=i, column=8, value=val)
        cell.border = border_all
        cell.number_format = "#,##0"

    # Bảng đăng ký phiên (tên sheet ngày) — thêm dòng khi có sheet mới
    ws["A14"] = "Danh sách phiên (tên sheet)"
    ws["A14"].font = sub_font
    ws["A15"] = "Tên_sheet"
    ws["A15"].font = hdr_font
    ws["A15"].fill = hdr_fill
    ws["B15"] = "Ghi chú"
    ws["B15"].font = hdr_font
    ws["B15"].fill = hdr_fill
    for c in ("A15", "B15"):
        ws[c].border = border_all

    # Ví dụ 1 dòng phiên
    ws["A16"] = SAMPLE_DAY
    ws["B16"] = "Xóa dòng mẫu hoặc giữ khi dùng sheet mẫu"

    tab_sessions = Table(displayName="tblPhien", ref="A15:B201")
    tab_sessions.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab_sessions)

    # Bảng công nợ
    start_row = 18
    ws.cell(row=start_row, column=1, value="Theo dõi công nợ").font = sub_font
    hdr_row = start_row + 1
    headers = ["Tên", "Nhóm", "Lũy kế phải đóng", "Đã thanh toán", "Số dư"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border_all

    data_start = hdr_row + 1
    fixed_labels = [f"Thành viên {i}" for i in range(1, 8)]
    for i, name in enumerate(fixed_labels):
        r = data_start + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value="Cố định")
        ws.cell(row=r, column=4, value=0)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border_all
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"

    gl_example_row = data_start + 7
    ws.cell(row=gl_example_row, column=1, value="(Thêm tên GL)")
    ws.cell(row=gl_example_row, column=2, value="GL")
    ws.cell(row=gl_example_row, column=4, value=0)
    for c in range(1, 6):
        ws.cell(row=gl_example_row, column=c).border = border_all
    ws.cell(row=gl_example_row, column=4).number_format = "#,##0"
    ws.cell(row=gl_example_row, column=5).number_format = "#,##0"

    last_table_row = 55
    tab_debt = Table(
        displayName="tblCongNo",
        ref=f"A{hdr_row}:E{last_table_row}",
    )
    tab_debt.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab_debt)

    for r in range(data_start, last_table_row + 1):
        ws.cell(row=r, column=3, value=luy_ke_cell_formula(r))
        ws.cell(row=r, column=5, value=so_du_cell_formula(r))
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"

    for r in range(gl_example_row + 1, last_table_row + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border_all
        ws.cell(row=r, column=4).number_format = "#,##0"

    ws["A50"] = "Tên TV cố định (tham chiếu sheet phiên)"
    ws["A50"].font = sub_font
    for i in range(7):
        r = FIXED_NAMES_START_ROW + i
        table_row = data_start + i
        ws.cell(row=r, column=1, value=f"=A{table_row}")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 18

    red_fill = PatternFill("solid", fgColor="F8CBAD")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    debt_col = f"E{data_start}:E{last_table_row}"
    ws.conditional_formatting.add(
        debt_col,
        CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        debt_col,
        CellIsRule(operator="lessThan", formula=["0"], fill=green_fill),
    )


def day_sheet_layout(ws, wb: Workbook, sheet_title: str, include_sample_gl: bool) -> None:
    mref = f"'{MASTER}'"
    pr = PARAM_EXTRA_COURT_ROW
    rate_extra = f"{mref}!$H${pr}"
    rate_gl_m = f"{mref}!$H${PARAM_GL_M_ROW}"
    rate_gl_f = f"{mref}!$H${PARAM_GL_F_ROW}"
    thr_hi = f"{mref}!$H${PARAM_THR_HIGH_ROW}"
    thr_lo = f"{mref}!$H${PARAM_THR_LOW_ROW}"

    ws["A1"] = f"Phiên ngày: {sheet_title}"
    ws["A1"].font = title_font

    ws["A3"] = "Nhập liệu buổi"
    ws["A3"].font = sub_font
    labels = [
        ("A4", "Tiền sân cố định (buổi)"),
        ("A5", "Số sân thêm"),
        ("A6", "Tiền cầu (shuttlecock)"),
        ("A7", "Tổng chi buổi (tự động)"),
        ("A8", "Số TV cố định có mặt"),
        ("A9", "Số GL"),
        ("A10", "Tổng người có mặt"),
        ("A11", "Tổng phí GL (khi > ngưỡng trên)"),
    ]
    for addr, lab in labels:
        ws[addr] = lab

    ws["B4"].value = 0
    ws["B5"].value = 0
    ws["B6"].value = 0
    ws["B7"].value = f"=B4+B5*{rate_extra}+B6"
    ws["B8"].value = "=COUNTIF(E20:E26,1)"
    ws["B9"].value = '=COUNTIFS(C28:C35,"GL",B28:B35,"<>")'
    ws["B10"].value = "=B8+B9"
    ws["B11"].value = (
        f"=IF(B10>{thr_hi},"
        f'COUNTIFS(C28:C35,"GL",D28:D35,"Nam",B28:B35,"<>")*{rate_gl_m}+'
        f'COUNTIFS(C28:C35,"GL",D28:D35,"Nữ",B28:B35,"<>")*{rate_gl_f},0)'
    )

    for addr in ("B4", "B5", "B6", "B7", "B11"):
        ws[addr].number_format = "#,##0"
    for addr in ("B8", "B9", "B10"):
        ws[addr].number_format = "0"

    ws["D4"] = "Kiểm tra tổng"
    ws["E4"] = '=IF(ABS(B7-SUM(F20:F35))<1,"Khớp tổng","Lệch tổng — kiểm tra")'
    ws["D5"] = "Cảnh báo"
    ws["E5"] = f"=IF(B10<{thr_lo},\"Dưới ngưỡng ít người — nên không đặt sân thêm\",\"\")"

    ws["A13"] = "Danh sách tham gia"
    ws["A13"].font = sub_font

    # Header bảng người
    hrow = 14
    heads = ["STT", "Tên", "Loại", "Giới tính (Nam/Nữ)", "Có mặt (1=có)", "Thành tiền (VNĐ)"]
    for col, h in enumerate(heads, start=1):
        cell = ws.cell(row=hrow, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border_all

    # 7 TV cố định
    first_data = 20
    for i in range(7):
        r = first_data + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=f"={mref}!$A${FIXED_NAMES_START_ROW + i}")
        ws.cell(row=r, column=3, value="Cố định")
        ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5, value=1)
        fcell = ws.cell(row=r, column=6)
        # Công thức thành tiền TV cố định
        fcell.value = (
            f"=IF(E{r}<>1,0,"
            f"IF($B$10=0,0,"
            f"IF($B$10>{thr_hi},"
            f"IF($B$8=0,0,MAX(0,($B$7-$B$11)/$B$8)),"
            f"$B$7/$B$10)))"
        )
        fcell.number_format = "#,##0"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border_all

    # GL rows 28-35
    gl_start = 28
    for i in range(8):
        r = gl_start + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=3, value="GL")
        ws.cell(row=r, column=2, value="")
        ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5, value="")
        fcell = ws.cell(row=r, column=6)
        fcell.value = (
            f'=IF(B{r}="",0,IF($B$10=0,0,IF($B$10>{thr_hi},'
            f'IF(D{r}="Nam",{rate_gl_m},IF(D{r}="Nữ",{rate_gl_f},0)),'
            f"$B$7/$B$10)))"
        )
        fcell.number_format = "#,##0"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border_all

    if include_sample_gl:
        ws["B28"] = "Khách A"
        ws["D28"] = "Nam"

    # GL: validation Nam/Nữ
    dv = DataValidation(type="list", formula1='"Nam,Nữ"', allow_blank=True)
    dv.error = "Chọn Nam hoặc Nữ cho GL"
    ws.add_data_validation(dv)
    dv.add(f"D{gl_start}:D{gl_start + 7}")

    # TV: validation có mặt 0/1
    dv2 = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    ws.add_data_validation(dv2)
    dv2.add(f"E{first_data}:E{first_data + 6}")

    # Ghi chú ngưỡng
    ws["A37"] = (
        f"Ghi chú: Phí GL riêng khi tổng người > giá trị «Ngưỡng trên» tại sheet «{MASTER}». "
        "Khi ≤ ngưỡng trên: chia đều tổng chi cho mọi người có mặt."
    )
    ws["A37"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A37:F39")

    # CF: thành tiền >0
    ws.conditional_formatting.add(
        f"F{first_data}:F{35}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFF2CC")),
    )

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18


def add_template(wb: Workbook) -> None:
    ws = wb.create_sheet(TEMPLATE)
    day_sheet_layout(ws, wb, "Mẫu — nhân bản sheet này", include_sample_gl=False)
    ws["A1"] = "MẪU PHIÊN — Nhân bản (chuột phải tab) → Đổi tên theo yyyy-mm-dd → Thêm tên sheet vào bảng tblPhien"


def add_sample_day(wb: Workbook) -> None:
    ws = wb.create_sheet(SAMPLE_DAY)
    day_sheet_layout(ws, wb, SAMPLE_DAY, include_sample_gl=True)
    ws["B4"] = 400_000
    ws["B5"] = 1
    ws["B6"] = 120_000


def add_guide(wb: Workbook) -> None:
    ws = wb.create_sheet(GUIDE)
    text = f"""HƯỚNG DẪN SỬ DỤNG

1) Môi trường: Microsoft Excel (2016 trở lên) hoặc LibreOffice Calc. Cột «Lũy kế» dùng SUMPRODUCT+SUMIF+INDIRECT (không cần Excel 365).

2) Thiết lập ban đầu:
   - Vào «{MASTER}», sửa tên 7 thành viên trong bảng công nợ (cột Tên).
   - Chỉnh «Tham số»: đơn giá sân thêm, phí GL Nam/Nữ, hai ngưỡng người.

3) Tạo phiên mới:
   - Chuột phải tab «{TEMPLATE}» → Move or Copy → Create a copy.
   - Đổi tên sheet đúng định dạng ngày: yyyy-mm-dd (ví dụ 2026-04-22).
   - Trên «{MASTER}», bảng tblPhien: thêm một dòng, cột Tên_sheet = đúng tên sheet vừa tạo.

4) Nhập buổi:
   - Điền tiền sân cố định, số sân thêm, tiền cầu.
   - Đánh dấu TV cố định có mặt (1) / vắng (0).
   - Thêm GL: tên + giới tính Nam/Nữ.

5) Công nợ:
   - Cột «Lũy kế phải đóng» tự cộng dồn từ mọi sheet có trong tblPhien.
   - Nhập «Đã thanh toán» khi ai đó trả tiền.
   - «Số dư» dương (tô đỏ) = còn nợ; âm (tô xanh) = dư.

6) GL mới: thêm một dòng trong bảng công nợ (Nhóm = GL), điền đúng tên trùng với tên trên sheet phiên.

7) Xóa / đổi tên sheet: cập nhật lại cột Tên_sheet trong tblPhien cho khớp.
"""
    ws["A1"] = text
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100


def main() -> None:
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    add_master(wb)
    add_guide(wb)
    add_template(wb)
    add_sample_day(wb)

    out = "/home/linhpd/workspace/badmintonManager/QuanLyCauLong_Nhom7.xlsx"
    wb.save(out)
    print(f"Đã tạo: {out}")


if __name__ == "__main__":
    main()
